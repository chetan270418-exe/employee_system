from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from functools import wraps
from datetime import date, datetime
from io import BytesIO
import calendar
from app import db
from app.models.user import User, Department
from app.models.attendance import Attendance
from app.models.leave import LeaveRequest
from app.models.payroll import Payroll

reports_bp = Blueprint('reports', __name__)


def admin_or_hr(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ('admin', 'hr'):
            flash('Access denied.', 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated


@reports_bp.route('/')
@login_required
@admin_or_hr
def index():
    return render_template('reports/index.html')


@reports_bp.route('/attendance')
@login_required
@admin_or_hr
def attendance_report():
    month = request.args.get('month', date.today().month, type=int)
    year = request.args.get('year', date.today().year, type=int)
    dept_id = request.args.get('dept_id', type=int)
    employees_q = User.query.filter_by(is_active=True, role='employee')
    if dept_id:
        employees_q = employees_q.filter_by(department_id=dept_id)
    employees = employees_q.all()
    working_days = sum(1 for d in range(1, calendar.monthrange(year, month)[1]+1)
                       if date(year, month, d).weekday() < 5)
    report_data = []
    for emp in employees:
        records = Attendance.query.filter(
            Attendance.employee_id == emp.id,
            db.extract('month', Attendance.date) == month,
            db.extract('year', Attendance.date) == year
        ).all()
        present = sum(1 for r in records if r.status == 'present')
        late = sum(1 for r in records if r.status == 'late')
        half = sum(1 for r in records if r.status == 'half_day')
        absent = working_days - present - late - half
        report_data.append({
            'emp': emp,
            'present': present, 'late': late,
            'half_day': half, 'absent': absent,
            'percentage': round((present + late) / working_days * 100, 1) if working_days else 0,
            'total_hours': round(sum(r.work_hours or 0 for r in records), 1),
        })
    departments = Department.query.all()
    return render_template('reports/attendance_report.html',
                           report_data=report_data, month=month, year=year,
                           working_days=working_days, departments=departments, dept_id=dept_id,
                           month_name=calendar.month_name[month])


@reports_bp.route('/payroll')
@login_required
@admin_or_hr
def payroll_report():
    month = request.args.get('month', date.today().month, type=int)
    year = request.args.get('year', date.today().year, type=int)
    payrolls = Payroll.query.filter_by(month=month, year=year).all()
    total_gross = sum(float(p.gross_salary) for p in payrolls)
    total_net = sum(float(p.net_salary) for p in payrolls)
    total_deductions = sum(float(p.total_deductions) for p in payrolls)
    return render_template('reports/payroll_report.html',
                           payrolls=payrolls, month=month, year=year,
                           month_name=calendar.month_name[month],
                           total_gross=total_gross, total_net=total_net,
                           total_deductions=total_deductions)


@reports_bp.route('/leave')
@login_required
@admin_or_hr
def leave_report():
    year = request.args.get('year', date.today().year, type=int)
    status = request.args.get('status', 'all')
    q = LeaveRequest.query.filter(
        db.extract('year', LeaveRequest.from_date) == year
    )
    if status != 'all':
        q = q.filter_by(status=status)
    leaves = q.order_by(LeaveRequest.from_date.desc()).all()
    return render_template('reports/leave_report.html', leaves=leaves, year=year, status=status)


@reports_bp.route('/export/attendance-excel')
@login_required
@admin_or_hr
def export_attendance_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    month = request.args.get('month', date.today().month, type=int)
    year = request.args.get('year', date.today().year, type=int)
    employees = User.query.filter_by(is_active=True, role='employee').all()
    working_days = sum(1 for d in range(1, calendar.monthrange(year, month)[1]+1)
                       if date(year, month, d).weekday() < 5)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Attendance {calendar.month_name[month]} {year}'
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    headers = ['Emp ID', 'Name', 'Department', 'Designation', 'Working Days',
               'Present', 'Late', 'Half Day', 'Absent', 'Attendance %', 'Total Hours']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14
    for row, emp in enumerate(employees, 2):
        records = Attendance.query.filter(
            Attendance.employee_id == emp.id,
            db.extract('month', Attendance.date) == month,
            db.extract('year', Attendance.date) == year
        ).all()
        present = sum(1 for r in records if r.status == 'present')
        late = sum(1 for r in records if r.status == 'late')
        half = sum(1 for r in records if r.status == 'half_day')
        absent = working_days - present - late - half
        pct = round((present + late) / working_days * 100, 1) if working_days else 0
        hours = round(sum(r.work_hours or 0 for r in records), 1)
        ws.append([emp.employee_id, emp.name,
                   emp.department.name if emp.department else '',
                   emp.designation or '', working_days,
                   present, late, half, absent, pct, hours])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'attendance_{calendar.month_abbr[month]}_{year}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@reports_bp.route('/export/payroll-excel')
@login_required
@admin_or_hr
def export_payroll_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    month = request.args.get('month', date.today().month, type=int)
    year = request.args.get('year', date.today().year, type=int)
    payrolls = Payroll.query.filter_by(month=month, year=year).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Payroll {calendar.month_name[month]} {year}'
    header_fill = PatternFill('solid', fgColor='1B5E20')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    headers = ['Emp ID', 'Name', 'Dept', 'Present Days', 'Base', 'HRA', 'DA', 'TA',
               'Overtime', 'Bonus', 'Gross', 'PF', 'ESI', 'TDS', 'Leave Ded', 'Total Ded', 'Net Salary', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14
    for p in payrolls:
        ws.append([p.employee.employee_id, p.employee.name,
                   p.employee.department.name if p.employee.department else '',
                   p.present_days, float(p.base_salary), float(p.hra), float(p.da),
                   float(p.ta), float(p.overtime_amount), float(p.bonus),
                   float(p.gross_salary), float(p.pf), float(p.esi), float(p.tds),
                   float(p.leave_deduction), float(p.total_deductions),
                   float(p.net_salary), p.status])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'payroll_{calendar.month_abbr[month]}_{year}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
